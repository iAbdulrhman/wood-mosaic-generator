"""Tests for manufacturing CSV and Excel exports."""

from __future__ import annotations

from io import BytesIO

import numpy as np
import pytest
from openpyxl import load_workbook
from PIL import Image

from modules.color_matcher import quantize_grid_image
from modules.cutlist_exporter import (
    PIECE_HEADERS,
    create_manufacturing_workbook,
    manufacturing_workbook_to_bytes,
    piece_records_to_csv_bytes,
    validate_export_inputs,
)
from modules.grid_calculator import calculate_grid
from modules.piece_generator import generate_piece_records


def create_export_results():
    """Create compatible grid, color, and piece results."""

    grid = calculate_grid(
        board_width_mm=41,
        board_height_mm=41,
        piece_size_mm=20,
        gap_mm=1,
    )

    pixels = np.array(
        [
            [
                [0, 0, 0],
                [255, 255, 255],
            ],
            [
                [255, 255, 255],
                [0, 0, 0],
            ],
        ],
        dtype=np.uint8,
    )

    image = Image.fromarray(pixels)

    colors = quantize_grid_image(
        grid_image=image,
        color_count=2,
        preview_cell_size=10,
    )

    records = generate_piece_records(
        grid=grid,
        colors=colors,
    )

    return grid, colors, records


def test_csv_contains_header_and_all_records():
    """Export one CSV row for every piece."""

    _, _, records = create_export_results()

    csv_bytes = piece_records_to_csv_bytes(
        records
    )

    text = csv_bytes.decode("utf-8-sig")
    lines = text.strip().splitlines()

    assert lines[0].split(",") == PIECE_HEADERS
    assert len(lines) == 5


def test_csv_contains_piece_identifiers():
    """Include piece identifiers in CSV output."""

    _, _, records = create_export_results()

    text = piece_records_to_csv_bytes(
        records
    ).decode("utf-8-sig")

    assert "R001-C001" in text
    assert "R002-C002" in text


def test_csv_rejects_empty_records():
    """Reject empty manufacturing CSV exports."""

    with pytest.raises(
        ValueError,
        match="At least one",
    ):
        piece_records_to_csv_bytes(())


def test_workbook_contains_required_sheets():
    """Create the three required worksheets."""

    grid, colors, records = create_export_results()

    workbook = create_manufacturing_workbook(
        grid=grid,
        colors=colors,
        records=records,
    )

    assert workbook.sheetnames == [
        "Project Info",
        "Pieces",
        "Color Summary",
    ]


def test_pieces_sheet_contains_all_records():
    """Write every piece to the Pieces worksheet."""

    grid, colors, records = create_export_results()

    workbook = create_manufacturing_workbook(
        grid=grid,
        colors=colors,
        records=records,
    )

    worksheet = workbook["Pieces"]

    assert worksheet.max_row == 5
    assert worksheet.max_column == len(
        PIECE_HEADERS
    )

    assert worksheet["A2"].value == (
        "R001-C001"
    )

    assert worksheet["A5"].value == (
        "R002-C002"
    )


def test_color_summary_matches_piece_total():
    """Ensure color quantities equal total piece count."""

    grid, colors, records = create_export_results()

    workbook = create_manufacturing_workbook(
        grid=grid,
        colors=colors,
        records=records,
    )

    worksheet = workbook["Color Summary"]

    piece_total = sum(
        worksheet.cell(
            row=row,
            column=6,
        ).value
        for row in range(
            2,
            worksheet.max_row + 1,
        )
    )

    assert piece_total == grid.total_pieces


def test_project_info_contains_dimensions():
    """Write board dimensions to the workbook."""

    grid, colors, records = create_export_results()

    workbook = create_manufacturing_workbook(
        grid=grid,
        colors=colors,
        records=records,
    )

    worksheet = workbook["Project Info"]

    values = {
        worksheet.cell(
            row=row,
            column=1,
        ).value: worksheet.cell(
            row=row,
            column=2,
        ).value
        for row in range(
            2,
            worksheet.max_row + 1,
        )
    }

    assert values["Board width"] == 41
    assert values["Board height"] == 41
    assert values["Total pieces"] == 4
    assert values["Actual colors"] == 2


def test_xlsx_bytes_can_be_reopened():
    """Create a valid downloadable XLSX file."""

    grid, colors, records = create_export_results()

    xlsx_bytes = manufacturing_workbook_to_bytes(
        grid=grid,
        colors=colors,
        records=records,
    )

    assert xlsx_bytes.startswith(b"PK")

    workbook = load_workbook(
        BytesIO(xlsx_bytes)
    )

    assert workbook.sheetnames == [
        "Project Info",
        "Pieces",
        "Color Summary",
    ]


def test_export_rejects_empty_records():
    """Reject an export without piece records."""

    grid, colors, _ = create_export_results()

    with pytest.raises(
        ValueError,
        match="At least one",
    ):
        validate_export_inputs(
            grid=grid,
            colors=colors,
            records=(),
        )
