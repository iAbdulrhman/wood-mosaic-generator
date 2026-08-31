"""Export wooden mosaic manufacturing records to CSV and Excel."""

from __future__ import annotations

import csv
from io import BytesIO, StringIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from modules.color_matcher import ColorQuantizationResult
from modules.grid_calculator import GridResult
from modules.piece_generator import PieceRecord


PIECE_HEADERS = [
    "piece_id",
    "sequence",
    "row",
    "column",
    "color_id",
    "hex_color",
    "red",
    "green",
    "blue",
    "x_mm",
    "y_mm",
    "center_x_mm",
    "center_y_mm",
    "width_mm",
    "height_mm",
]


def validate_export_inputs(
    grid: GridResult,
    colors: ColorQuantizationResult,
    records: tuple[PieceRecord, ...],
) -> None:
    """Validate data before creating manufacturing exports."""

    if not records:
        raise ValueError(
            "At least one piece record is required."
        )

    if len(records) != grid.total_pieces:
        raise ValueError(
            "Piece record count does not match the physical grid."
        )

    if len(records) != colors.total_pieces:
        raise ValueError(
            "Piece record count does not match the color grid."
        )

    if grid.columns != colors.columns:
        raise ValueError(
            "Grid columns do not match the color result."
        )

    if grid.rows != colors.rows:
        raise ValueError(
            "Grid rows do not match the color result."
        )


def piece_record_to_row(
    record: PieceRecord,
) -> list:
    """Convert one piece record to an ordered export row."""

    return [
        record.piece_id,
        record.sequence,
        record.row,
        record.column,
        record.color_id,
        record.hex_color,
        record.red,
        record.green,
        record.blue,
        record.x_mm,
        record.y_mm,
        record.center_x_mm,
        record.center_y_mm,
        record.width_mm,
        record.height_mm,
    ]


def piece_records_to_csv_bytes(
    records: tuple[PieceRecord, ...],
) -> bytes:
    """Export piece records as UTF-8 CSV bytes."""

    if not records:
        raise ValueError(
            "At least one piece record is required."
        )

    text_buffer = StringIO()

    writer = csv.writer(
        text_buffer,
        lineterminator="\n",
    )

    writer.writerow(PIECE_HEADERS)

    for record in records:
        writer.writerow(
            piece_record_to_row(record)
        )

    return text_buffer.getvalue().encode(
        "utf-8-sig"
    )


def _style_header_row(worksheet) -> None:
    """Apply common styling to the first worksheet row."""

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="704628",
    )

    header_font = Font(
        color="FFFFFF",
        bold=True,
    )

    thin_border = Border(
        bottom=Side(
            style="thin",
            color="D9C8B8",
        )
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        cell.border = thin_border

    worksheet.row_dimensions[1].height = 24


def _set_column_widths(
    worksheet,
    minimum_width: int = 10,
    maximum_width: int = 30,
) -> None:
    """Set readable worksheet column widths."""

    for column_cells in worksheet.columns:
        column_index = column_cells[0].column
        column_letter = get_column_letter(
            column_index
        )

        longest_value = 0

        for cell in column_cells:
            if cell.value is None:
                continue

            longest_value = max(
                longest_value,
                len(str(cell.value)),
            )

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max(
                longest_value + 2,
                minimum_width,
            ),
            maximum_width,
        )


def _create_project_info_sheet(
    workbook: Workbook,
    grid: GridResult,
    colors: ColorQuantizationResult,
) -> None:
    """Create the project-information worksheet."""

    worksheet = workbook.create_sheet(
        title="Project Info"
    )

    worksheet.append(
        [
            "Property",
            "Value",
            "Unit",
        ]
    )

    rows = [
        (
            "Board width",
            grid.board_width_mm,
            "mm",
        ),
        (
            "Board height",
            grid.board_height_mm,
            "mm",
        ),
        (
            "Piece size",
            grid.piece_size_mm,
            "mm",
        ),
        (
            "Gap",
            grid.gap_mm,
            "mm",
        ),
        (
            "Columns",
            grid.columns,
            "",
        ),
        (
            "Rows",
            grid.rows,
            "",
        ),
        (
            "Total pieces",
            grid.total_pieces,
            "pieces",
        ),
        (
            "Used width",
            grid.used_width_mm,
            "mm",
        ),
        (
            "Used height",
            grid.used_height_mm,
            "mm",
        ),
        (
            "Horizontal margin",
            grid.margin_x_mm,
            "mm",
        ),
        (
            "Vertical margin",
            grid.margin_y_mm,
            "mm",
        ),
        (
            "Requested colors",
            colors.requested_color_count,
            "colors",
        ),
        (
            "Actual colors",
            colors.actual_color_count,
            "colors",
        ),
    ]

    for row in rows:
        worksheet.append(row)

    _style_header_row(worksheet)
    _set_column_widths(worksheet)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def _create_pieces_sheet(
    workbook: Workbook,
    records: tuple[PieceRecord, ...],
) -> None:
    """Create the complete piece cut-list worksheet."""

    worksheet = workbook.create_sheet(
        title="Pieces"
    )

    worksheet.append(PIECE_HEADERS)

    color_fills: dict[str, PatternFill] = {}

    for record in records:
        worksheet.append(
            piece_record_to_row(record)
        )

        if record.hex_color not in color_fills:
            color_fills[record.hex_color] = PatternFill(
                fill_type="solid",
                fgColor=record.hex_color.lstrip("#"),
            )

        color_cell = worksheet.cell(
            row=worksheet.max_row,
            column=6,
        )

        color_cell.fill = color_fills[
            record.hex_color
        ]

    _style_header_row(worksheet)
    _set_column_widths(worksheet)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False


def _create_color_summary_sheet(
    workbook: Workbook,
    colors: ColorQuantizationResult,
) -> None:
    """Create a color-quantities worksheet."""

    worksheet = workbook.create_sheet(
        title="Color Summary"
    )

    worksheet.append(
        [
            "color_id",
            "hex_color",
            "red",
            "green",
            "blue",
            "piece_count",
            "percentage",
        ]
    )

    for entry in colors.palette:
        red, green, blue = entry.rgb

        worksheet.append(
            [
                entry.color_id,
                entry.hex_color,
                red,
                green,
                blue,
                entry.piece_count,
                entry.percentage / 100,
            ]
        )

        hex_cell = worksheet.cell(
            row=worksheet.max_row,
            column=2,
        )

        hex_cell.fill = PatternFill(
            fill_type="solid",
            fgColor=entry.hex_color.lstrip("#"),
        )

        percentage_cell = worksheet.cell(
            row=worksheet.max_row,
            column=7,
        )

        percentage_cell.number_format = "0.00%"

    _style_header_row(worksheet)
    _set_column_widths(worksheet)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def create_manufacturing_workbook(
    grid: GridResult,
    colors: ColorQuantizationResult,
    records: tuple[PieceRecord, ...],
) -> Workbook:
    """Create a formatted manufacturing Excel workbook."""

    validate_export_inputs(
        grid=grid,
        colors=colors,
        records=records,
    )

    workbook = Workbook()

    default_sheet = workbook.active
    workbook.remove(default_sheet)

    _create_project_info_sheet(
        workbook=workbook,
        grid=grid,
        colors=colors,
    )

    _create_pieces_sheet(
        workbook=workbook,
        records=records,
    )

    _create_color_summary_sheet(
        workbook=workbook,
        colors=colors,
    )

    workbook.properties.title = (
        "Wood Mosaic Manufacturing List"
    )

    workbook.properties.subject = (
        "Piece positions, colors, and quantities"
    )

    workbook.properties.creator = (
        "Wood Mosaic Generator"
    )

    return workbook


def manufacturing_workbook_to_bytes(
    grid: GridResult,
    colors: ColorQuantizationResult,
    records: tuple[PieceRecord, ...],
) -> bytes:
    """Create downloadable XLSX bytes."""

    workbook = create_manufacturing_workbook(
        grid=grid,
        colors=colors,
        records=records,
    )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output.getvalue()
